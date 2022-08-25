from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Geburtstage"


ws.append(['Name', 'Geburtsdatum', 'Spitznamen', 'Stadt'])


ws['A2'].value = "Junus"
ws['B2'].value = "03.10.1991"

ws['A3'].value = "Hans Müller"
ws['B3'].value = "01.10.1993"

# ws.delete_rows(2)
ws.move_range("A1:D3", rows=3, cols=2, translate=True)
wb.save('Geburtstage.xlsx')