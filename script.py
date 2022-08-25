from openpyxl import Workbook, load_workbook

wb = load_workbook("Bestellungen.xlsx")
ws = wb["Bestellungen"]

def get_price(name):
    preisliste = wb["Preisliste"]
    for row in range(2, 12):
        if preisliste["a" + str(row)].value == name:
            return preisliste["b" + str(row)].value
    return 9999999

# Preise in Tabelle ausfüllen
for row in range(2,23):
    product_name = ws["b"  + str(row)].value
    ws["d"  + str(row)].value = get_price(product_name)


# Gesamtpreis für Bestellungen berechnen
for row in range(2,23):
    c = ws["c" + str(row)].value # Werte von c2, c3, ..., c22
    d = ws["d" + str(row)].value # d2, d3, ..., d22
    result = c * d
    ws["e"  + str(row)].value = result

wb.save("Bestellungen.xlsx")
